# -*- coding: utf-8 -*-
"""

write cases for huge excel files

needed:
    csv reader
    row counter

need heavy rework to make this publishable


"""

import os
import time
from openpyxl import load_workbook, Workbook
import matplotlib.pyplot as plt
import db_management as dbm


def main(excel=True,plot=True):
    
    
    #name = "TIC_2016-11-01T1212"
    name = "TIC_MAST_3_7"
    inputfilename = "../input/%s.xlsx"%name
    inputsheetname = "%s.csv"%name
    
    
    WBI = load_workbook(inputfilename)
    data_sheet_ranges = WBI[inputsheetname]    
 
    """
    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.patch.set_facecolor('black')
    """
    if excel:   
        pass
        """
        # outdated
        outputfilename = "../output/Trim-%s.xlsx"%name
        outputsheetname = inputsheetname
        
        WBO = Workbook()
        output_sheet_ranges = WBO.active
        output_sheet_ranges.title=outputsheetname

        cot = 1
        output_sheet_ranges['A%d'%cot].value = "id"
        output_sheet_ranges['B%d'%cot].value = "ra"
        output_sheet_ranges['C%d'%cot].value = "dec"
        output_sheet_ranges['D%d'%cot].value = "Tmag"
        output_sheet_ranges['E%d'%cot].value = "e_Tmag"
        output_sheet_ranges['F%d'%cot].value = "Teff"
        output_sheet_ranges['G%d'%cot].value = "e_Teff"
        output_sheet_ranges['H%d'%cot].value = "type"
        """

    else:
        remove = True
        
        db_name = "../input/TIC_db_MAST.db"
        table_name = "Data"
        if not os.path.isfile(db_name):
            c,conn = dbm.create_db(db_name, remove)
            columns = [("ID","text"),("TMAG","FLOAT"),("TEFF","FLOAT"),("PMRA","FLOAT"),
                       ("PMDEC","FLOAT"),("RA","FLOAT"),("DEC","FLOAT"),("OBJTYPE","text")]
            primary = "ID"
            control = dbm.create_table_prime(table_name,primary,columns,conn,c)
        else:
            c,conn = dbm.access_db(db_name)
        
        


    
    # need a fast way of figuring out the number of data point fast
    for i in range(100):
        num = i+6
        cot = i+2
        
        id = data_sheet_ranges['A%d'%num].value
        if id == "" or id == None:
            continue
        
        ra = data_sheet_ranges['L%d'%num].value
        dec = data_sheet_ranges['M%d'%num].value
        Tmag = data_sheet_ranges['BD%d'%num].value
        e_Tmag = data_sheet_ranges['BE%d'%num].value
        Teff = data_sheet_ranges['BH%d'%num].value
        e_Teff = data_sheet_ranges['BI%d'%num].value
        
        if Teff == "NaN" or Teff == "" or Teff == None:
            type = "galaxy"
        else:
            type = "star"

        if excel:
            output_sheet_ranges['A%d'%cot].value = id
            output_sheet_ranges['B%d'%cot].value = ra
            output_sheet_ranges['C%d'%cot].value = dec
            output_sheet_ranges['D%d'%cot].value = Tmag
            output_sheet_ranges['E%d'%cot].value = e_Tmag
            output_sheet_ranges['F%d'%cot].value = Teff
            output_sheet_ranges['G%d'%cot].value = e_Teff
            output_sheet_ranges['H%d'%cot].value = type
        
        else:
            # database functions
            data = [id,ra,dec,Tmag,e_Tmag,Teff,e_Teff,type]
            dbm.insert_single_data(table_name, data,conn,c)
        
        if plot:
            plt.plot(ra,dec,"o",color="w",markersize=(20-Tmag))

    if excel:
        WBO.save(outputfilename)
        
    if plot:
        plt.xlabel("ra")
        plt.ylabel("dec")
        plt.show()

def main2(excel=True,plot=True):
    
    
    #name = "TIC_2016-11-01T1212"
    
    start = time.time()
    name = "TIC_MAST_3_7"
    inputfilename = "../../input/%s_trim.xlsx"%name
    inputsheetname = "%s.csv"%name
    
    
    WBI = load_workbook(inputfilename)
    data_sheet_ranges = WBI[inputsheetname]    


    # outdated
    outputfilename = "../output/Trim-%s.xlsx"%name
    outputsheetname = inputsheetname
    
    WBO = Workbook()
    output_sheet_ranges = WBO.active
    output_sheet_ranges.title=outputsheetname

    cot = 1
    output_sheet_ranges['A%d'%cot].value = "ID"
    output_sheet_ranges['B%d'%cot].value = "UCACID"
    output_sheet_ranges['C%d'%cot].value = "RA"
    output_sheet_ranges['D%d'%cot].value = "DEC"
    output_sheet_ranges['E%d'%cot].value = "PMRA"
    output_sheet_ranges['F%d'%cot].value = "PMDEC"
    output_sheet_ranges['G%d'%cot].value = "VMAG"
    output_sheet_ranges['H%d'%cot].value = "JMAG"
    output_sheet_ranges['I%d'%cot].value = "TMAG"
    output_sheet_ranges['J%d'%cot].value = "TEFF"        
    
    
    print time.time()-start
    start = time.time()
    count = 0
    for i in range(86000):
        
        
        num = i+6
        
    
        ID = data_sheet_ranges['A%d'%num].value
        UCAC4ID = data_sheet_ranges['B%d'%num].value
        
        if ID == "" or ID == None:
            continue
        
        if UCAC4ID == "" or UCAC4ID == None:
            continue
        
        
        ra = data_sheet_ranges['C%d'%num].value
        dec = data_sheet_ranges['D%d'%num].value
        
        if ra > 310 or ra < 290:
            continue
        if dec > 57 or dec < 43:
            continue
        
        cot +=1
        pmra = data_sheet_ranges['E%d'%num].value
        pmdec = data_sheet_ranges['F%d'%num].value
        Vmag = data_sheet_ranges['G%d'%num].value
        Jmag = data_sheet_ranges['H%d'%num].value
        Tmag = data_sheet_ranges['I%d'%num].value
        Teff = data_sheet_ranges['J%d'%num].value
            
        output_sheet_ranges['A%d'%cot].value = ID
        output_sheet_ranges['B%d'%cot].value = UCAC4ID
        output_sheet_ranges['C%d'%cot].value = ra
        output_sheet_ranges['D%d'%cot].value = dec
        output_sheet_ranges['E%d'%cot].value = pmra
        output_sheet_ranges['F%d'%cot].value = pmdec
        output_sheet_ranges['G%d'%cot].value = Vmag
        output_sheet_ranges['H%d'%cot].value = Jmag
        output_sheet_ranges['I%d'%cot].value = Tmag
        output_sheet_ranges['J%d'%cot].value = Teff    
         
        count +=1
        
    WBO.save("../../input/TIC_MAST_3_7_actual.xlsx")
    print time.time()-start
    print count
    
if __name__ == "__main__":
    main2(False,False)
    
    
    """
    db_name = "../database/TIC_db_MAST.db"
    table_name = "Data"    
    c,conn = dbm.access_db(db_name)
    data = dbm.return_all_from_table(table_name,conn,c)
    for i in data:
        print i
    """
    
    
    
    